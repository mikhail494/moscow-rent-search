from datetime import datetime
from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import load_workbook

from app.models import Listing
from app.services.export import (
    EXCEL_HEADERS,
    ExportSearch,
    NoExportResultsError,
    create_excel_export,
    create_html_export,
    displayed_listings,
)


class ExportTests(unittest.TestCase):
    def setUp(self) -> None:
        self.output = TemporaryDirectory()
        self.created_at = datetime(2026, 7, 17, 12, 30, 45)
        self.search = ExportSearch(
            source="test",
            property_types=("studio", "one_room"),
            min_area=18,
            max_area=50,
            max_price=90000,
            listings=(
                Listing(
                    source="test",
                    external_id="confirmed",
                    url="https://example.test/confirmed",
                    title="Студия",
                    property_type="studio",
                    rent_price=50000,
                    area_sqm=24.5,
                    metro_station="Тверская",
                    metro_minutes=6,
                    address="Москва, Тверская улица, 1",
                    latitude=55.76,
                    longitude=37.61,
                    location_verified=True,
                ),
                Listing(
                    source="test",
                    external_id="unverified",
                    url="https://example.test/unverified",
                    title="Однокомнатная квартира",
                    property_type="one_room",
                    rent_price=60000,
                    area_sqm=38,
                    metro_station=None,
                    metro_minutes=None,
                    address="Москва, улица Пример, 2",
                    latitude=None,
                    longitude=None,
                    location_verified=False,
                ),
            ),
        )

    def tearDown(self) -> None:
        self.output.cleanup()

    @property
    def output_dir(self) -> Path:
        return Path(self.output.name)

    def test_creates_readable_excel_with_links_and_expected_rows(self) -> None:
        export_path = create_excel_export(
            self.search, self.output_dir, True, self.created_at
        )

        self.assertEqual(export_path.name, "rent_search_2026-07-17_12-30-45.xlsx")
        workbook = load_workbook(export_path)
        worksheet = workbook.active
        self.assertEqual(tuple(cell.value for cell in worksheet[1]), EXCEL_HEADERS)
        self.assertEqual(worksheet.max_row, 3)
        self.assertEqual(worksheet.freeze_panes, "A2")
        self.assertEqual(worksheet.auto_filter.ref, "A1:K3")
        self.assertEqual(worksheet["I2"].hyperlink.target, "https://example.test/confirmed")
        self.assertEqual(worksheet["A2"].number_format, "#,##0")
        self.assertEqual(worksheet["C2"].number_format, "0.0")
        self.assertTrue(worksheet["A1"].font.bold)

    def test_creates_self_contained_html_with_filters_and_links(self) -> None:
        export_path = create_html_export(
            self.search, self.output_dir, True, self.created_at
        )

        html = export_path.read_text(encoding="utf-8")
        self.assertEqual(export_path.name, "rent_search_2026-07-17_12-30-45.html")
        self.assertIn("Применённые фильтры", html)
        self.assertIn("2026-07-17 12:30:45", html)
        self.assertIn('href="https://example.test/confirmed"', html)
        self.assertNotIn("/static/", html)
        self.assertEqual(html.count("<tbody>"), 1)
        self.assertEqual(html.count("<tr>"), 3)

    def test_raises_clear_error_without_results(self) -> None:
        with self.assertRaisesRegex(NoExportResultsError, "Нет результатов для экспорта."):
            create_excel_export(None, self.output_dir, True, self.created_at)

    def test_excludes_unverified_rows_when_they_are_hidden_in_table(self) -> None:
        listings = displayed_listings(self.search, include_unverified=False)

        self.assertEqual([listing.external_id for listing in listings], ["confirmed"])


if __name__ == "__main__":
    unittest.main()
