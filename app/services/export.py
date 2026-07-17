from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from html import escape
from pathlib import Path
from typing import Literal, Sequence

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models import Listing


PropertyType = Literal["studio", "one_room"]

EXCEL_HEADERS = (
    "Цена",
    "Тип квартиры",
    "Площадь",
    "Метро",
    "Время до метро",
    "Адрес",
    "Статус местоположения",
    "Источник",
    "Ссылка",
    "Latitude",
    "Longitude",
)

SOURCE_LABELS = {
    "test": "Тестовые данные",
    "cian": "ЦИАН",
    "yandex_realty": "Яндекс Недвижимость",
}


class NoExportResultsError(ValueError):
    """Raised when the current search has no rows suitable for export."""


@dataclass(frozen=True)
class ExportSearch:
    source: str
    property_types: tuple[PropertyType, ...]
    min_area: float | None
    max_area: float | None
    max_price: int
    listings: tuple[Listing, ...]


def displayed_listings(
    search: ExportSearch | None, include_unverified: bool
) -> list[Listing]:
    if search is None:
        return []

    listings = [
        listing
        for listing in search.listings
        if include_unverified or listing.location_verified
    ]
    return sorted(
        listings,
        key=lambda listing: (not listing.location_verified, listing.rent_price),
    )


def create_excel_export(
    search: ExportSearch | None,
    output_dir: Path,
    include_unverified: bool,
    created_at: datetime | None = None,
) -> Path:
    listings = _require_listings(search, include_unverified)
    timestamp = created_at or datetime.now()
    output_path = _output_path(output_dir, timestamp, "xlsx")

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Результаты"
    worksheet.append(EXCEL_HEADERS)

    header_fill = PatternFill("solid", fgColor="176D5D")
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    for listing in listings:
        worksheet.append(
            (
                listing.rent_price,
                property_type_label(listing.property_type),
                listing.area_sqm,
                listing.metro_station or "",
                listing.metro_minutes,
                listing.address,
                location_status_label(listing),
                source_label(listing.source),
                listing.url,
                listing.latitude,
                listing.longitude,
            )
        )
        link_cell = worksheet.cell(row=worksheet.max_row, column=9)
        link_cell.hyperlink = listing.url
        link_cell.style = "Hyperlink"

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:K{worksheet.max_row}"
    for row_index in range(2, worksheet.max_row + 1):
        worksheet.cell(row=row_index, column=1).number_format = "#,##0"
        worksheet.cell(row=row_index, column=3).number_format = "0.0"

    for column_index, width in enumerate(
        (14, 22, 12, 22, 18, 42, 34, 24, 46, 14, 14), start=1
    ):
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    workbook.save(output_path)
    return output_path


def create_html_export(
    search: ExportSearch | None,
    output_dir: Path,
    include_unverified: bool,
    created_at: datetime | None = None,
) -> Path:
    listings = _require_listings(search, include_unverified)
    assert search is not None
    timestamp = created_at or datetime.now()
    output_path = _output_path(output_dir, timestamp, "html")

    filters = (
        ("Источник", source_label(search.source)),
        (
            "Тип квартиры",
            ", ".join(property_type_label(item) for item in search.property_types),
        ),
        ("Минимальная площадь", area_filter_label(search.min_area)),
        ("Максимальная площадь", area_filter_label(search.max_area)),
        ("Максимальная цена", f"{search.max_price:,} ₽".replace(",", " ")),
        (
            "Объявления без подтверждённых координат",
            "Показаны" if include_unverified else "Скрыты",
        ),
    )
    filter_rows = "\n".join(
        f"<dt>{escape(label)}</dt><dd>{escape(value)}</dd>" for label, value in filters
    )
    listing_rows = "\n".join(_html_listing_row(listing) for listing in listings)
    created_label = timestamp.strftime("%Y-%m-%d %H:%M:%S")

    output_path.write_text(
        f"""<!doctype html>
<html lang="ru">
<head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>Результаты поиска аренды в Москве</title>
    <style>
        body {{ margin: 32px; color: #17231f; background: #f8faf8; font: 15px Arial, sans-serif; }}
        main {{ max-width: 1440px; margin: 0 auto; }}
        h1 {{ margin: 0 0 6px; }}
        .created {{ margin: 0 0 28px; color: #5d6a66; }}
        dl {{ display: grid; grid-template-columns: max-content 1fr; gap: 8px 18px; margin: 0 0 28px; }}
        dt {{ color: #5d6a66; font-weight: 700; }}
        dd {{ margin: 0; }}
        table {{ width: 100%; border-collapse: collapse; background: #fff; }}
        th, td {{ padding: 12px; border: 1px solid #c9d3cf; text-align: left; vertical-align: top; }}
        th {{ color: #fff; background: #176d5d; }}
        a {{ color: #176d5d; font-weight: 700; }}
    </style>
</head>
<body>
    <main>
        <h1>Результаты поиска аренды в Москве</h1>
        <p class="created">Создано: {escape(created_label)}</p>
        <h2>Применённые фильтры</h2>
        <dl>{filter_rows}</dl>
        <h2>Объявления ({len(listings)})</h2>
        <table>
            <thead>
                <tr>
                    {''.join(f'<th>{escape(header)}</th>' for header in EXCEL_HEADERS)}
                </tr>
            </thead>
            <tbody>
                {listing_rows}
            </tbody>
        </table>
    </main>
</body>
</html>
""",
        encoding="utf-8",
    )
    return output_path


def property_type_label(property_type: PropertyType) -> str:
    return "Студия" if property_type == "studio" else "Однокомнатная квартира"


def source_label(source: str) -> str:
    return SOURCE_LABELS.get(source, source)


def location_status_label(listing: Listing) -> str:
    return (
        "В области, координаты подтверждены"
        if listing.location_verified
        else "Местоположение не подтверждено"
    )


def _require_listings(
    search: ExportSearch | None, include_unverified: bool
) -> list[Listing]:
    listings = displayed_listings(search, include_unverified)
    if not listings:
        raise NoExportResultsError("Нет результатов для экспорта.")
    return listings


def _output_path(output_dir: Path, timestamp: datetime, extension: str) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    filename = f"rent_search_{timestamp.strftime('%Y-%m-%d_%H-%M-%S')}.{extension}"
    return output_dir / filename


def area_filter_label(value: float | None) -> str:
    if value is None:
        return "Не задана"
    return f"{value:g} м²"


def _html_listing_row(listing: Listing) -> str:
    url = escape(listing.url, quote=True)
    values: Sequence[str] = (
        f"{listing.rent_price:,} ₽".replace(",", " "),
        property_type_label(listing.property_type),
        f"{listing.area_sqm:g} м²",
        listing.metro_station or "—",
        f"{listing.metro_minutes} мин" if listing.metro_minutes is not None else "—",
        listing.address,
        location_status_label(listing),
        source_label(listing.source),
        "",
        "" if listing.latitude is None else f"{listing.latitude:g}",
        "" if listing.longitude is None else f"{listing.longitude:g}",
    )
    cells = [f"<td>{escape(value)}</td>" for value in values[:8]]
    cells.append(f'<td><a href="{url}">Открыть</a></td>')
    cells.extend(f"<td>{escape(value)}</td>" for value in values[9:])
    return f"<tr>{''.join(cells)}</tr>"
